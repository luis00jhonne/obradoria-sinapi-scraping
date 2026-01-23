import pandas as pd
import requests
import os
import glob
import re
import unicodedata
from openpyxl import load_workbook
from datetime import datetime

# ------------------------------
# CONFIGURAÇÕES
# ------------------------------
API_BASE = "http://localhost:8891"
# Produção: API_BASE = "https://api.obradoria.com.br"
COMPOSICAO_API = f"{API_BASE}/api/composicoes/lote"

# Token JWT da API. Obtenha em POST /api/auth/login e exporte antes de rodar:
#   export OBRADORIA_TOKEN="..."
BEARER_TOKEN = os.getenv("OBRADORIA_TOKEN", "")
AUTH_HEADERS = {
    'Content-Type': 'application/json',
    'Authorization': f'Bearer {BEARER_TOKEN}'
}

if not BEARER_TOKEN:
    print("⚠️  OBRADORIA_TOKEN não definido — a API responderá 401.")

# Diretório das planilhas, relativo a este script (baixar do dataset no Zenodo)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "excels")

# Arquivos de log, também relativos ao script (não ao diretório de execução)
LOG_FILE = os.path.join(BASE_DIR, "log_envio_composicoes.csv")
ERRO_FILE = os.path.join(BASE_DIR, "erros_envio.csv")

def remover_acentos(texto):
    if texto is None or not isinstance(texto, str):
        return str(texto) if texto is not None else ''
    return ''.join(c for c in unicodedata.normalize('NFD', texto)
                   if unicodedata.category(c) != 'Mn')

def extrair_codigo_hiperlink(formula):
    """Extrai o código do segundo argumento da função HIPERLIGAÇÃO/HYPERLINK."""
    if not formula or not isinstance(formula, str):
        return None

    # A fórmula pode ser:
    # PT: =HIPERLIGAÇÃO("#"&CÉL("address";DESLOCAMENTO(...));103370)
    # EN: =HYPERLINK("#"&CELL("address",OFFSET(...)),105003)
    # O segundo argumento é o código que queremos

    # Buscar o último número antes do último parêntese de fechamento
    # Aceita tanto vírgula (EN) quanto ponto e vírgula (PT) como separador
    match = re.search(r'[,;](\d+)\)\s*$', formula)
    if match:
        return int(match.group(1))

    return None

def ler_excel(file_path):
    """Lê o Excel e retorna DataFrame, extraindo códigos de fórmulas HYPERLINK."""
    skiprows = 9
    sheet_name = "CSD"

    # Carregar workbook duas vezes: uma para fórmulas e outra para valores
    # 1. Para valores gerais (data_only=True)
    wb_data = load_workbook(filename=file_path, data_only=True, read_only=True)
    ws_data = wb_data[sheet_name]

    # 2. Para fórmulas da coluna Código da Composição (data_only=False)
    wb_formulas = load_workbook(filename=file_path, data_only=False, read_only=True)
    ws_formulas = wb_formulas[sheet_name]

    # Extrair todas as linhas com valores (OTIMIZADO: uma única leitura)
    all_rows = list(ws_data.iter_rows(values_only=True))

    # Extrair todas as linhas com fórmulas (OTIMIZADO: uma única leitura)
    all_formula_rows = list(ws_formulas.iter_rows(values_only=False))

    # Pegar o cabeçalho (linha na posição skiprows)
    cols = all_rows[skiprows] if skiprows < len(all_rows) else None

    # Pegar os dados (linhas após o cabeçalho)
    vals = all_rows[skiprows+1:] if skiprows+1 < len(all_rows) else []

    # Criar DataFrame
    df = pd.DataFrame(vals, columns=cols)

    # Encontrar a coluna "Código da Composição" no cabeçalho original
    codigo_col_idx = None
    for idx, col in enumerate(cols):
        if col and 'Composição' in str(col) and 'digo' in str(col):
            codigo_col_idx = idx
            break

    # Substituir valores da coluna "Código da Composição" lendo das fórmulas
    if codigo_col_idx is not None:
        codigos_extraidos = []

        # Processar exatamente len(df) linhas
        num_linhas_processar = min(len(df), len(all_formula_rows) - skiprows - 1, len(all_rows) - skiprows - 1)

        # Iterar sobre as linhas de fórmulas correspondentes aos dados
        for i in range(num_linhas_processar):
            row_idx = skiprows + 1 + i

            # Verificar se estamos dentro dos limites
            if row_idx >= len(all_formula_rows):
                codigos_extraidos.append(None)
                continue

            cell = all_formula_rows[row_idx][codigo_col_idx]
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                # É uma fórmula, extrair o código
                codigo = extrair_codigo_hiperlink(cell.value)
                codigos_extraidos.append(codigo)
            else:
                # Não é fórmula, usar o valor direto de all_rows
                valor_direto = all_rows[row_idx][codigo_col_idx] if row_idx < len(all_rows) else None
                codigos_extraidos.append(valor_direto)

        # Atualizar a coluna no DataFrame
        if len(codigos_extraidos) == len(df):
            # Converter para Series com dtype que aceita nulos (Int64 com I maiúsculo)
            df.iloc[:, codigo_col_idx] = pd.Series(codigos_extraidos, dtype='Int64')

    wb_data.close()
    wb_formulas.close()

    # Limpar colunas principais
    df.columns = df.columns.str.strip()
    df.columns = df.columns.str.replace('\n', ' ')
    df.columns = [remover_acentos(col) for col in df.columns]

    return df

def carregar_log(clear=False):
    """Carrega log de envios para evitar duplicação.

    Se clear for True, retorna DataFrame vazio (limpa o histórico para cada arquivo).
    """
    if clear or not os.path.exists(LOG_FILE):
        return pd.DataFrame(columns=['codigo', 'data_envio'])
    return pd.read_csv(LOG_FILE, dtype={'codigo': str})

def salvar_log(log_df):
    """Salva log atualizado."""
    log_df.to_csv(LOG_FILE, index=False)

def registrar_erro(codigos, erro):
    """Registra erros em arquivo separado."""
    erro_df = pd.DataFrame({
        'codigos': [','.join(codigos)],
        'erro': [erro],
        'data': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    })
    
    if os.path.exists(ERRO_FILE):
        erro_df.to_csv(ERRO_FILE, mode='a', header=False, index=False)
    else:
        erro_df.to_csv(ERRO_FILE, index=False)

def enviar_em_lote(df, log_envio, batch_size=500):
    """Envia composições em lote, pulando já enviadas."""
    
    # Preparar códigos e filtrar já enviados
    df['codigo_str'] = df['Codigo da Composicao'].astype(str).str.strip()
    
    codigos_enviados = set(log_envio['codigo'].astype(str)) if not log_envio.empty else set()
    df_novos = df[~df['codigo_str'].isin(codigos_enviados)]
    
    print(f"  📊 Total: {len(df)} | Já enviados: {len(codigos_enviados)} | A enviar: {len(df_novos)}")
    
    if df_novos.empty:
        print("  ✓ Todas as composições já foram enviadas!")
        return log_envio, 0, 0
    
    # Preparar payloads
    payloads = []
    codigos = []
    
    for _, row in df_novos.iterrows():
        codigo = row['codigo_str']
        payload = {
            "grupo": row["Grupo"],
            "codigo": codigo,
            "nome": row["Descricao"],
            "unidadeMedida": row["Unidade"]
        }
        payloads.append(payload)
        codigos.append(codigo)
    
    # Enviar em lotes
    total = len(payloads)
    enviados = []
    total_sucesso = 0
    total_falha = 0
    
    for i in range(0, total, batch_size):
        batch = payloads[i:i + batch_size]
        batch_codigos = codigos[i:i + batch_size]
        
        try:
            # Log de request completo para debug via Postman / API
            import json
            print(f"  ▶️ ENVIANDO batch {i//batch_size + 1}/{(total + batch_size - 1)//batch_size}")
            print(f"    URL: {COMPOSICAO_API}")
            print(f"    Autenticado: {'sim' if BEARER_TOKEN else 'NÃO'}")
            print(f"    Payload items: {len(batch)}")
            print(f"    Payload amostra (JSON): {json.dumps(batch[:2], ensure_ascii=False, indent=2)}")

            response = requests.post(
                COMPOSICAO_API,
                json=batch,
                headers=AUTH_HEADERS,
                timeout=60
            )

            print(f"  📨 HTTP {response.status_code} | batch {i//batch_size + 1} | {len(batch)} itms")
            print(f"    Response headers: {dict(response.headers)}")
            print(f"    Response body: {response.text[:1000]}" if response.text else "    Response body: <vazio>")

            if response.status_code in [200, 201, 202]:
                enviados.extend(batch_codigos)
                total_sucesso += len(batch)
                print(f"  ✓ Lote {i//batch_size + 1}/{(total + batch_size - 1)//batch_size}: {len(batch)} enviados")
            else:
                total_falha += len(batch)
                registrar_erro(batch_codigos, f"HTTP {response.status_code}: {response.text[:200]}")
                print(f"  ✗ Erro lote {i//batch_size + 1}: {response.status_code}")
                
        except Exception as e:
            total_falha += len(batch)
            registrar_erro(batch_codigos, str(e))
            print(f"  ✗ Erro conexão lote {i//batch_size + 1}: {e}")
    
    # Atualizar log com códigos enviados com sucesso
    if enviados:
        novos_logs = pd.DataFrame({
            'codigo': enviados,
            'data_envio': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        log_envio = pd.concat([log_envio, novos_logs], ignore_index=True)
        salvar_log(log_envio)
    
    return log_envio, total_sucesso, total_falha

def main():
    print("=" * 60)
    print("INICIANDO PROCESSAMENTO DE COMPOSIÇÕES")
    print("=" * 60)
    
    all_files = glob.glob(os.path.join(EXCEL_PATH, "*.xlsx"))
    
    if not all_files:
        print("❌ Nenhum arquivo Excel encontrado em:", EXCEL_PATH)
        return

    print(f"\n📁 Encontrados {len(all_files)} arquivo(s)\n")
    
    total_geral_sucesso = 0
    total_geral_falha = 0

    for idx, file in enumerate(all_files, 1):
        nome_arquivo = os.path.basename(file)
        print(f"\n[{idx}/{len(all_files)}] 📄 Processando: {nome_arquivo}")

        try:
            df = ler_excel(file)
            # Limpar log a cada novo arquivo para processar códigos independentes por arquivo
            log_envio = carregar_log(clear=True)
            log_envio, sucessos, falhas = enviar_em_lote(df, log_envio)

            total_geral_sucesso += sucessos
            total_geral_falha += falhas

            print(f"  ✓ Concluído: {sucessos} sucessos, {falhas} falhas")

        except Exception as e:
            print(f"  ✗ Erro ao processar arquivo: {e}")
            continue

    print("\n" + "=" * 60)
    print("RESUMO FINAL")
    print("=" * 60)
    print(f"✓ Total enviado com sucesso: {total_geral_sucesso}")
    print(f"✗ Total com falha: {total_geral_falha}")
    print(f"📊 Total processado: {total_geral_sucesso + total_geral_falha}")

    if total_geral_falha > 0:
        print(f"⚠️  Verifique erros em: {ERRO_FILE}")

    print("=" * 60)

if __name__ == "__main__":
    main()