import pandas as pd
import requests
import os
import glob
import re
import unicodedata
from openpyxl import load_workbook
from datetime import datetime
from decimal import Decimal, InvalidOperation

# ------------------------------
# CONFIGURAÇÕES
# ------------------------------
API_BASE = "http://localhost:8891"
# Produção: API_BASE = "https://api.obradoria.com.br"
PRECO_COMPOSICAO_API_LOTE = f"{API_BASE}/api/preco-composicoes/lote"

# Token JWT da API. Obtenha em POST /api/auth/login e exporte antes de rodar:
#   export OBRADORIA_TOKEN="..."
BEARER_TOKEN = os.getenv("OBRADORIA_TOKEN", "")
AUTH_HEADERS = {
    'Content-Type': 'application/json',
    'Authorization': f'Bearer {BEARER_TOKEN}'
}

if not BEARER_TOKEN:
    print("⚠️  OBRADORIA_TOKEN não definido — a API responderá 401.")

# Diretório das planilhas, relativo a este script (baixar do dataset no Zenodo)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "excels")

# Arquivos de log, também relativos ao script (não ao diretório de execução)
LOG_FILE = os.path.join(BASE_DIR, "log_envio_precos_composicoes.csv")
ERRO_FILE = os.path.join(BASE_DIR, "erros_envio_precos_composicoes.csv")

# ------------------------------
# FUNÇÕES AUXILIARES
# ------------------------------

def remover_acentos(texto):
    if texto is None or not isinstance(texto, str):
        return str(texto) if texto is not None else ''
    return ''.join(c for c in unicodedata.normalize('NFD', texto)
                   if unicodedata.category(c) != 'Mn')


def to_decimal(value):
    """Converte valores para Decimal ou None (BigDecimal compatível)."""
    if value is None:
        return None

    if isinstance(value, str):
        value = value.strip().replace(',', '.')
        if value == '' or value.upper() == 'NAN':
            return None

    if pd.isna(value):
        return None

    try:
        # Evita precisão de float indesejada
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return None

def ler_dados_cabecalho(file_path, sheet_name):
    """Lê células B3 e B4 para pegar mês de referência e data de emissão."""
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb[sheet_name]
    
    mes_referencia = ws['B3'].value  # Ex: "09/2025"
    data_emissao = ws['B4'].value    # Ex: "10/10/2025"
    
    # Converter data de emissão para formato ISO
    if isinstance(data_emissao, str):
        try:
            # Tentar converter de DD/MM/YYYY
            data_emissao = datetime.strptime(data_emissao, '%d/%m/%Y').strftime('%Y-%m-%d')
        except:
            data_emissao = datetime.now().strftime('%Y-%m-%d')
    elif hasattr(data_emissao, 'strftime'):
        data_emissao = data_emissao.strftime('%Y-%m-%d')
    else:
        data_emissao = datetime.now().strftime('%Y-%m-%d')
    
    wb.close()
    return mes_referencia, data_emissao

def extrair_codigo_hiperlink(formula):
    """Extrai o código do segundo argumento da função HIPERLIGAÇÃO/HYPERLINK."""
    if not formula or not isinstance(formula, str):
        return None

    # A fórmula pode ser:
    # PT: =HIPERLIGAÇÃO("#"&CÉL("address";DESLOCAMENTO(...));103370)
    # EN: =HYPERLINK("#"&CELL("address",OFFSET(...)),105003)
    # O segundo argumento é o código que queremos

    # Buscar o último número antes do último parêntese de fechamento
    # Aceita tanto vírgula (EN) quanto ponto e vírgula (PT) como separador
    match = re.search(r'[,;](\d+)\)\s*$', formula)
    if match:
        return int(match.group(1))

    # Se não encontrou, tentar debug
    print(f"DEBUG: Não conseguiu extrair código da fórmula: {formula}")
    return None

def ler_precos_composicoes_excel(file_path, sheet_name, skiprows=9):
    """Lê preços de composições do Excel a partir da linha especificada."""
    # Carregar workbook duas vezes: uma para fórmulas e outra para valores
    # 1. Para valores gerais (data_only=True)
    wb_data = load_workbook(filename=file_path, data_only=True, read_only=True)
    ws_data = wb_data[sheet_name]

    # 2. Para fórmulas da coluna Código da Composição (data_only=False)
    wb_formulas = load_workbook(filename=file_path, data_only=False, read_only=True)
    ws_formulas = wb_formulas[sheet_name]

    # Extrair todas as linhas com valores (OTIMIZADO: uma única leitura)
    all_rows = list(ws_data.iter_rows(values_only=True))

    # Extrair todas as linhas com fórmulas (OTIMIZADO: uma única leitura)
    all_formula_rows = list(ws_formulas.iter_rows(values_only=False))

    # Pegar o cabeçalho (linha na posição skiprows)
    cols = all_rows[skiprows] if skiprows < len(all_rows) else None

    # Pegar os dados (linhas após o cabeçalho)
    vals = all_rows[skiprows+1:] if skiprows+1 < len(all_rows) else []

    # Criar DataFrame
    df = pd.DataFrame(vals, columns=cols)

    print(f"DEBUG: DataFrame criado com {len(df)} linhas e {len(df.columns)} colunas")
    print(f"DEBUG: len(all_rows)={len(all_rows)}, len(all_formula_rows)={len(all_formula_rows)}, skiprows={skiprows}")

    # Encontrar a coluna "Código da Composição" no cabeçalho original
    codigo_col_idx = None
    for idx, col in enumerate(cols):
        if col and 'Composição' in str(col) and 'digo' in str(col):
            codigo_col_idx = idx
            print(f"DEBUG: Coluna 'Código da Composição' encontrada no índice {idx}")
            break

    # Substituir valores da coluna "Código da Composição" lendo das fórmulas
    if codigo_col_idx is not None:
        codigos_extraidos = []

        # Processar exatamente len(df) linhas
        num_linhas_processar = min(len(df), len(all_formula_rows) - skiprows - 1, len(all_rows) - skiprows - 1)
        print(f"DEBUG: Processando {num_linhas_processar} linhas de códigos")

        # Iterar sobre as linhas de fórmulas correspondentes aos dados
        for i in range(num_linhas_processar):
            row_idx = skiprows + 1 + i

            # Verificar se estamos dentro dos limites
            if row_idx >= len(all_formula_rows):
                print(f"DEBUG: AVISO - row_idx {row_idx} >= len(all_formula_rows) {len(all_formula_rows)}")
                codigos_extraidos.append(None)
                continue

            cell = all_formula_rows[row_idx][codigo_col_idx]
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                # É uma fórmula, extrair o código
                if i < 3:  # Debug: mostrar as primeiras 3 fórmulas
                    print(f"DEBUG: Fórmula linha {i}: {cell.value[:150]}")
                codigo = extrair_codigo_hiperlink(cell.value)
                codigos_extraidos.append(codigo)
                if i < 3:  # Debug: mostrar os primeiros 3 códigos extraídos
                    print(f"DEBUG: Código extraído: {codigo}")
            else:
                # Não é fórmula, usar o valor direto de all_rows
                valor_direto = all_rows[row_idx][codigo_col_idx] if row_idx < len(all_rows) else None
                codigos_extraidos.append(valor_direto)
                if i < 3:  # Debug
                    print(f"DEBUG: Valor direto linha {i}: {valor_direto}")

        print(f"DEBUG: Total de códigos extraídos: {len(codigos_extraidos)}, len(df): {len(df)}")

        # Atualizar a coluna no DataFrame
        if len(codigos_extraidos) == len(df):
            # Converter para Series com dtype que aceita nulos (Int64 com I maiúsculo)
            df.iloc[:, codigo_col_idx] = pd.Series(codigos_extraidos, dtype='Int64')
            print(f"DEBUG: ✓ Coluna atualizada com {len(codigos_extraidos)} códigos")
            print(f"DEBUG: Primeiros 10 códigos: {codigos_extraidos[:10]}")
            print(f"DEBUG: Códigos None: {sum(1 for c in codigos_extraidos if c is None)}")
        else:
            print(f"DEBUG: ✗ ERRO - Tamanhos não batem! len(codigos_extraidos)={len(codigos_extraidos)} != len(df)={len(df)}")
            print(f"DEBUG: Não foi possível atualizar a coluna 'Código da Composição'")

    # Ler também a linha de cabeçalho que contém as UFs (linha 9 do Excel = índice 8)
    header_ufs = list(all_rows[8]) if len(all_rows) > 8 else []

    wb_data.close()
    wb_formulas.close()

    print(df.iloc[:3, 1])

    # Limpar colunas principais
    df.columns = df.columns.str.strip()
    df.columns = df.columns.str.replace('\n', ' ')
    df.columns = [remover_acentos(col) for col in df.columns]

    print(df.iloc[:3, 1])

    # Detectar colunas de UF a partir do cabeçalho extra
    uf_cols = []
    uf_start_idx = None

    # Encontrar onde começam as UFs (geralmente depois de "Unidade")
    for idx, col_name in enumerate(df.columns):
        if 'Unidade' in col_name:
            uf_start_idx = idx + 1
            break

    if uf_start_idx:
        print(f"DEBUG: uf_start_idx={uf_start_idx}, len(header_ufs)={len(header_ufs)}, len(df.columns)={len(df.columns)}")

        # Percorrer colunas após "Unidade" e capturar UFs
        col_idx = uf_start_idx
        while col_idx < len(header_ufs):
            # Parar se ultrapassou o tamanho do df
            if col_idx >= len(df.columns):
                print(f"DEBUG: Parando loop - col_idx ({col_idx}) >= len(df.columns) ({len(df.columns)})")
                break

            uf = str(header_ufs[col_idx]).strip()

            # Verificar se é uma UF válida (2 letras maiúsculas)
            if len(uf) == 2 and uf.isupper() and uf.isalpha():
                print(f"DEBUG: UF encontrada: {uf} na posição col_idx={col_idx}")
                uf_cols.append(uf)

                # Assumir que as próximas 2 colunas são Custo e AC
                if col_idx < len(df.columns):
                    df.columns.values[col_idx] = f"{uf}_Custo"
                if col_idx + 1 < len(df.columns):
                    df.columns.values[col_idx + 1] = f"{uf}_AC"

                col_idx += 2  # Pular Custo e AC
            else:
                col_idx += 1

    # Converter colunas de preço para float
    for uf in uf_cols:
        if f"{uf}_Custo" in df.columns:
            df[f"{uf}_Custo"] = df[f"{uf}_Custo"].astype(str).str.replace(',', '.').str.strip()
            df[f"{uf}_Custo"] = pd.to_numeric(df[f"{uf}_Custo"], errors='coerce')
        
        if f"{uf}_AC" in df.columns:
            df[f"{uf}_AC"] = df[f"{uf}_AC"].astype(str).str.replace(',', '.').str.replace('%', '').str.strip()
            df[f"{uf}_AC"] = pd.to_numeric(df[f"{uf}_AC"], errors='coerce')
    
    return df, uf_cols

def carregar_log(arquivo):
    """Carrega log de envios."""
    if os.path.exists(arquivo):
        return pd.read_csv(arquivo, dtype={'chave': str})
    return pd.DataFrame(columns=['chave', 'data_envio'])

def salvar_log(log_df, arquivo):
    """Salva log atualizado."""
    log_df.to_csv(arquivo, index=False)

def registrar_erro(chaves, erro):
    """Registra erros em arquivo separado."""
    erro_df = pd.DataFrame({
        'chaves': [','.join(map(str, chaves))],
        'erro': [erro[:500]],
        'data': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    })
    
    if os.path.exists(ERRO_FILE):
        erro_df.to_csv(ERRO_FILE, mode='a', header=False, index=False)
    else:
        erro_df.to_csv(ERRO_FILE, index=False)

# ------------------------------
# PROCESSAMENTO DOS PREÇOS DE COMPOSIÇÕES
# ------------------------------

def processar_precos_composicoes(df_sem_desoneracao, df_com_desoneracao, uf_cols, 
                                  mes_referencia, data_emissao):
    """Processa DataFrames e extrai preços de composições por UF."""
    precos = []
    chaves = []
    
    # Criar dicionário de preços COM desoneração por código
    precos_com_desoneracao = {}
    if df_com_desoneracao is not None:
        df_com_desoneracao['codigo_str'] = df_com_desoneracao['Codigo da Composicao'].astype(str).str.strip()
        for _, row in df_com_desoneracao.iterrows():
            codigo = row['codigo_str']
            precos_com_desoneracao[codigo] = row
    
    # Processar DataFrame SEM desoneração (principal)
    df_sem_desoneracao['codigo_str'] = df_sem_desoneracao['Codigo da Composicao'].astype(str).str.strip()
    
    for _, row_sem in df_sem_desoneracao.iterrows():
        try:
            codigo = int(float(row_sem['codigo_str']))
        except:
            continue
        
        codigo_str = row_sem['codigo_str']
        
        # Buscar linha correspondente COM desoneração
        row_com = precos_com_desoneracao.get(codigo_str)
        
        # Processar cada UF
        for uf in uf_cols:
            # Custo e AC SEM desoneração
            custo_sem = to_decimal(row_sem.get(f"{uf}_Custo"))
            ac_sem = to_decimal(row_sem.get(f"{uf}_AC"))

            # Custo e AC COM desoneração (se existir)
            custo_com = to_decimal(row_com.get(f"{uf}_Custo")) if row_com is not None else None
            ac_com = to_decimal(row_com.get(f"{uf}_AC")) if row_com is not None else None

            # Normalizar zero como None (opcional; se precisa enviar 0, remova essa seção)
            if custo_sem is not None and custo_sem == Decimal('0'):
                custo_sem = None
            if ac_sem is not None and ac_sem == Decimal('0'):
                ac_sem = None
            if custo_com is not None and custo_com == Decimal('0'):
                custo_com = None
            if ac_com is not None and ac_com == Decimal('0'):
                ac_com = None

            # Converter Decimal para float para JSON
            custo_sem = float(custo_sem) if custo_sem is not None else None
            ac_sem = float(ac_sem) if ac_sem is not None else None
            custo_com = float(custo_com) if custo_com is not None else None
            ac_com = float(ac_com) if ac_com is not None else None
            
            # Criar payload (sempre envia, mesmo com valores nulos)
            payload = {
                "codigoComposicao": codigo,
                "mesAnoReferencia": mes_referencia,
                "dataEmissao": data_emissao,
                "estado": uf,
                "custoSemDesoneracao": custo_sem,
                "acSemDesoneracao": ac_sem,
                "custoComDesoneracao": custo_com,
                "acComDesoneracao": ac_com
            }
            
            # Criar chave única (codigo|uf|mes_ref)
            chave = f"{codigo}|{uf}|{mes_referencia}"
            
            precos.append(payload)
            chaves.append(chave)
    
    return precos, chaves

# ------------------------------
# ENVIO EM LOTE
# ------------------------------

def enviar_precos_lote(precos, chaves, log_precos, batch_size=500):
    """Envia preços em lote, pulando já enviados."""
    
    # Filtrar já enviados
    chaves_enviadas = set(log_precos['chave'].astype(str)) if not log_precos.empty else set()
    
    precos_novos = []
    chaves_novas = []
    
    for preco, chave in zip(precos, chaves):
        if chave not in chaves_enviadas:
            precos_novos.append(preco)
            chaves_novas.append(chave)
    
    print(f"  💰 Total: {len(precos)} | Já enviados: {len(chaves_enviadas)} | A enviar: {len(precos_novos)}")
    
    if not precos_novos:
        print("  ✓ Todos os preços já foram enviados!")
        return log_precos, len(precos), 0
    
    # Enviar em lotes
    total = len(precos_novos)
    enviados = []
    total_sucesso = 0
    total_falha = 0
    
    for i in range(0, total, batch_size):
        batch = precos_novos[i:i + batch_size]
        batch_chaves = chaves_novas[i:i + batch_size]
        
        try:
            import json
            print(f"  ▶️ ENVIANDO lote {i//batch_size + 1}/{(total + batch_size - 1)//batch_size}")
            print(f"    URL: {PRECO_COMPOSICAO_API_LOTE}")
            print(f"    Autenticado: {'sim' if BEARER_TOKEN else 'NÃO'}")
            print(f"    Quantidade itens: {len(batch)}")
            print(f"    Payload amostra (JSON): {json.dumps(batch[:2], ensure_ascii=False, indent=2)}")

            response = requests.post(
                PRECO_COMPOSICAO_API_LOTE,
                json=batch,
                headers=AUTH_HEADERS,
                timeout=60
            )

            print(f"  📨 HTTP {response.status_code} | lote {i//batch_size + 1}/{(total + batch_size - 1)//batch_size} | {len(batch)} itens")
            print(f"    Response headers: {dict(response.headers)}")
            print(f"    Response body: {response.text[:1000]}" if response.text else "    Response body: <vazio>")
            
            if response.status_code in [200, 201, 202]:
                enviados.extend(batch_chaves)
                total_sucesso += len(batch)
                print(f"    ✓ Lote {i//batch_size + 1}/{(total + batch_size - 1)//batch_size}: {len(batch)} enviados")
            else:
                total_falha += len(batch)
                registrar_erro(batch_chaves, f"HTTP {response.status_code}: {response.text[:200]}")
                print(f"    ✗ Erro lote {i//batch_size + 1}: HTTP {response.status_code}")
                
        except Exception as e:
            total_falha += len(batch)
            registrar_erro(batch_chaves, str(e))
            print(f"    ✗ Erro conexão lote {i//batch_size + 1}: {e}")
    
    # Atualizar log
    if enviados:
        novos_logs = pd.DataFrame({
            'chave': enviados,
            'data_envio': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        log_precos = pd.concat([log_precos, novos_logs], ignore_index=True)
        salvar_log(log_precos, LOG_FILE)
    
    return log_precos, total_sucesso, total_falha

# ------------------------------
# PROCESSO PRINCIPAL
# ------------------------------

def main():
    print("=" * 70)
    print("INICIANDO PROCESSAMENTO DE PREÇOS DE COMPOSIÇÕES")
    print("=" * 70)
    
    log_precos = carregar_log(LOG_FILE)
    
    all_files = glob.glob(os.path.join(EXCEL_PATH, "*.xlsx"))
    
    if not all_files:
        print("❌ Nenhum arquivo Excel encontrado em:", EXCEL_PATH)
        return
    
    print(f"\n📁 Encontrados {len(all_files)} arquivo(s)\n")
    
    total_sucesso = 0
    total_falha = 0
    
    for idx, file in enumerate(all_files, 1):
        nome_arquivo = os.path.basename(file)
        print(f"\n[{idx}/{len(all_files)}] 📄 Processando: {nome_arquivo}")
        
        try:
            # Ler dados do cabeçalho (usa CSD como referência)
            mes_referencia, data_emissao = ler_dados_cabecalho(file, "CSD")
            print(f"  📅 Mês: {mes_referencia} | Data: {data_emissao}")
            
            # Ler planilha CSD (SEM desoneração)
            df_csd, uf_cols = ler_precos_composicoes_excel(file, sheet_name="CSD", skiprows=9)
            
            print(f"  📊 CSD: {len(df_csd)} linhas | UFs: {', '.join(uf_cols)}")
            
            # Ler planilha CCD (COM desoneração) - pode não existir
            df_ccd = None
            try:
                df_ccd, _ = ler_precos_composicoes_excel(file, sheet_name="CCD", skiprows=9)
                print(f"  📊 CCD: {len(df_ccd)} linhas")
            except Exception as e:
                print(f"  ⚠️  Planilha CCD não encontrada ou erro ao ler: {e}")

            # Processar preços
            precos, chaves = processar_precos_composicoes(
                df_csd, df_ccd, uf_cols,
                mes_referencia, data_emissao
            )
            
            # Enviar preços
            log_precos, sucessos, falhas = enviar_precos_lote(precos, chaves, log_precos)
            total_sucesso += sucessos
            total_falha += falhas
            
            print(f"  ✓ Arquivo concluído: {sucessos} sucessos, {falhas} falhas")
            
        except Exception as e:
            print(f"  ✗ Erro ao processar arquivo: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    print("\n" + "=" * 70)
    print("RESUMO FINAL")
    print("=" * 70)
    print(f"✓ Total enviado com sucesso: {total_sucesso}")
    print(f"✗ Total com falha: {total_falha}")
    print(f"📊 Total processado: {total_sucesso + total_falha}")
    
    if total_falha > 0:
        print(f"\n⚠️  Verifique erros em: {ERRO_FILE}")
    
    print("=" * 70)

if __name__ == "__main__":
    main()