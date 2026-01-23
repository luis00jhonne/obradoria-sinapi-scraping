import pandas as pd
import requests
import os
import glob
import unicodedata
from openpyxl import load_workbook
from datetime import datetime

# ------------------------------
# CONFIGURAÇÕES
# ------------------------------
API_BASE = "http://localhost:8891"
# Produção: API_BASE = "https://api.obradoria.com.br"
PRECO_INSUMO_API_LOTE = f"{API_BASE}/api/preco-insumos/lote"

# Token JWT da API. Obtenha em POST /api/auth/login e exporte antes de rodar:
#   export OBRADORIA_TOKEN="..."
BEARER_TOKEN = os.getenv("OBRADORIA_TOKEN", "")
AUTH_HEADERS = {
    'Content-Type': 'application/json',
    'Authorization': f'Bearer {BEARER_TOKEN}'
}

if not BEARER_TOKEN:
    print("⚠️  OBRADORIA_TOKEN não definido — a API responderá 401.")

# Diretório das planilhas, relativo a este script (baixar do dataset no Zenodo)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "excels")

# Arquivos de log, também relativos ao script (não ao diretório de execução)
LOG_FILE = os.path.join(BASE_DIR, "log_envio_precos_insumos.csv")
ERRO_FILE = os.path.join(BASE_DIR, "erros_envio_precos_insumos.csv")

# ------------------------------
# FUNÇÕES AUXILIARES
# ------------------------------

def remover_acentos(texto):
    return ''.join(c for c in unicodedata.normalize('NFD', texto)
                   if unicodedata.category(c) != 'Mn')

def ler_dados_cabecalho(file_path, sheet_name):
    """Lê células B3 e B4 para pegar mês de referência e data de emissão."""
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb[sheet_name]
    
    mes_referencia = ws['B3'].value  # Ex: "09/2025"
    data_emissao = ws['B4'].value    # Ex: "10/10/2025"
    
    # Converter data de emissão para formato ISO
    if isinstance(data_emissao, str):
        try:
            # Tentar converter de DD/MM/YYYY
            data_emissao = datetime.strptime(data_emissao, '%d/%m/%Y').strftime('%Y-%m-%d')
        except:
            data_emissao = datetime.now().strftime('%Y-%m-%d')
    elif hasattr(data_emissao, 'strftime'):
        data_emissao = data_emissao.strftime('%Y-%m-%d')
    else:
        data_emissao = datetime.now().strftime('%Y-%m-%d')
    
    wb.close()
    return mes_referencia, data_emissao

def ler_precos_excel(file_path, sheet_name, skiprows=9):
    """Lê preços do Excel a partir da linha especificada."""
    df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=skiprows)
    df.columns = df.columns.str.strip()
    df.columns = df.columns.str.replace('\n', ' ')
    df.columns = [remover_acentos(col) for col in df.columns]
    
    # Converter colunas de preço para float
    # Detectar colunas de UF (geralmente são siglas de 2 letras em maiúsculo)
    uf_cols = [col for col in df.columns if len(col) == 2 and col.isupper()]
    
    for uf in uf_cols:
        if uf in df.columns:
            # Substituir vírgula por ponto e converter para float
            df[uf] = df[uf].astype(str).str.replace(',', '.').str.strip()
            df[uf] = pd.to_numeric(df[uf], errors='coerce')
    
    return df, uf_cols

def carregar_log(arquivo):
    """Carrega log de envios."""
    if os.path.exists(arquivo):
        return pd.read_csv(arquivo, dtype={'chave': str})
    return pd.DataFrame(columns=['chave', 'mes_referencia', 'data_envio'])

def salvar_log(log_df, arquivo):
    """Salva log atualizado."""
    log_df.to_csv(arquivo, index=False)

def registrar_erro(chaves, erro):
    """Registra erros em arquivo separado."""
    erro_df = pd.DataFrame({
        'chaves': [','.join(map(str, chaves))],
        'erro': [erro[:500]],
        'data': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    })
    
    if os.path.exists(ERRO_FILE):
        erro_df.to_csv(ERRO_FILE, mode='a', header=False, index=False)
    else:
        erro_df.to_csv(ERRO_FILE, index=False)

# ------------------------------
# PROCESSAMENTO DOS PREÇOS
# ------------------------------

def processar_precos(df_sem_desoneracao, df_com_desoneracao, uf_cols, 
                     mes_referencia, data_emissao):
    """Processa DataFrames e extrai preços por UF."""
    precos = []
    chaves = []
    
    # Criar dicionário de preços COM desoneração por código
    precos_com_desoneracao = {}
    if df_com_desoneracao is not None:
        df_com_desoneracao['codigo_str'] = df_com_desoneracao['Codigo do Insumo'].astype(str).str.strip()
        for _, row in df_com_desoneracao.iterrows():
            codigo = row['codigo_str']
            precos_com_desoneracao[codigo] = row
    
    # Processar DataFrame SEM desoneração (principal)
    df_sem_desoneracao['codigo_str'] = df_sem_desoneracao['Codigo do Insumo'].astype(str).str.strip()
    
    for _, row_sem in df_sem_desoneracao.iterrows():
        try:
            codigo = int(float(row_sem['codigo_str']))
        except:
            continue
        
        codigo_str = row_sem['codigo_str']
        
        # Ler origem de preço da coluna
        origem = row_sem.get('Origem de Preco')
        if pd.isna(origem) or str(origem).strip() == '':
            origem = "CR"  # Padrão se não houver valor
        else:
            origem = str(origem).strip()
        
        # Buscar linha correspondente COM desoneração
        row_com = precos_com_desoneracao.get(codigo_str)
        
        # Processar cada UF
        for uf in uf_cols:
            # Preço SEM desoneração
            preco_sem = row_sem.get(uf)
            if pd.isna(preco_sem) or preco_sem == 0:
                continue
            
            # Preço COM desoneração (se existir)
            preco_com = None
            if row_com is not None:
                preco_com = row_com.get(uf)
                if pd.isna(preco_com) or preco_com == 0:
                    preco_com = None
            
            # Criar payload
            payload = {
                "codigoInsumo": codigo,
                "origemPreco": origem,
                "mesAnoReferencia": mes_referencia,
                "dataEmissao": data_emissao,
                "estado": uf,
                "coeficienteSemDesoneracao": float(preco_sem),
                "coeficienteComDesoneracao": float(preco_com) if preco_com else None
            }
            
            # Criar chave única (codigo|uf|mes_ref)
            chave = f"{codigo}|{uf}|{mes_referencia}"
            
            precos.append(payload)
            chaves.append(chave)
    
    return precos, chaves

# ------------------------------
# ENVIO EM LOTE
# ------------------------------

def enviar_precos_lote(precos, chaves, log_precos, mes_referencia, batch_size=500):
    """Envia preços em lote, pulando já enviados."""
    
    # Filtrar já enviados (considerando mês de referência)
    chaves_enviadas = set(log_precos['chave'].astype(str)) if not log_precos.empty else set()
    
    precos_novos = []
    chaves_novas = []
    
    for preco, chave in zip(precos, chaves):
        if chave not in chaves_enviadas:
            precos_novos.append(preco)
            chaves_novas.append(chave)
    
    print(f"  💰 Total: {len(precos)} | Já enviados: {len(chaves_enviadas)} | A enviar: {len(precos_novos)}")
    
    if not precos_novos:
        print("  ✓ Todos os preços já foram enviados!")
        return log_precos, len(precos), 0
    
    # Enviar em lotes
    total = len(precos_novos)
    enviados = []
    total_sucesso = 0
    total_falha = 0
    
    for i in range(0, total, batch_size):
        batch = precos_novos[i:i + batch_size]
        batch_chaves = chaves_novas[i:i + batch_size]
        
        try:
            response = requests.post(
                PRECO_INSUMO_API_LOTE,
                json=batch,
                headers=AUTH_HEADERS,
                timeout=60
            )
            
            if response.status_code in [200, 201, 202]:
                enviados.extend(batch_chaves)
                total_sucesso += len(batch)
                print(f"    ✓ Lote {i//batch_size + 1}/{(total + batch_size - 1)//batch_size}: {len(batch)} enviados")
            else:
                total_falha += len(batch)
                registrar_erro(batch_chaves, f"HTTP {response.status_code}: {response.text[:200]}")
                print(f"    ✗ Erro lote {i//batch_size + 1}: HTTP {response.status_code}")
                
        except Exception as e:
            total_falha += len(batch)
            registrar_erro(batch_chaves, str(e))
            print(f"    ✗ Erro conexão lote {i//batch_size + 1}: {e}")
    
    # Atualizar log com mês de referência
    if enviados:
        novos_logs = pd.DataFrame({
            'chave': enviados,
            'mes_referencia': mes_referencia,
            'data_envio': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        log_precos = pd.concat([log_precos, novos_logs], ignore_index=True)
        salvar_log(log_precos, LOG_FILE)
    
    return log_precos, total_sucesso, total_falha

# ------------------------------
# PROCESSO PRINCIPAL
# ------------------------------

def main():
    print("=" * 70)
    print("INICIANDO PROCESSAMENTO DE PREÇOS DE INSUMOS")
    print("=" * 70)
    
    log_precos = carregar_log(LOG_FILE)
    
    all_files = glob.glob(os.path.join(EXCEL_PATH, "*.xlsx"))
    
    if not all_files:
        print("❌ Nenhum arquivo Excel encontrado em:", EXCEL_PATH)
        return
    
    print(f"\n📁 Encontrados {len(all_files)} arquivo(s)\n")
    
    total_sucesso = 0
    total_falha = 0
    
    for idx, file in enumerate(all_files, 1):
        nome_arquivo = os.path.basename(file)
        print(f"\n[{idx}/{len(all_files)}] 📄 Processando: {nome_arquivo}")
        
        try:
            # Ler dados do cabeçalho
            mes_referencia, data_emissao = ler_dados_cabecalho(file, "ISD")
            print(f"  📅 Mês: {mes_referencia} | Data: {data_emissao}")
            
            # Ler planilha ISD (SEM desoneração)
            df_isd, uf_cols = ler_precos_excel(file, sheet_name="ISD", skiprows=9)
            print(f"  📊 ISD: {len(df_isd)} linhas | UFs: {', '.join(uf_cols)}")
            
            # Ler planilha ICD (COM desoneração) - pode não existir
            df_icd = None
            try:
                df_icd, _ = ler_precos_excel(file, sheet_name="ICD", skiprows=9)
                print(f"  📊 ICD: {len(df_icd)} linhas")
            except Exception as e:
                print(f"  ⚠️  Planilha ICD não encontrada ou erro ao ler: {e}")
            
            # Processar preços
            precos, chaves = processar_precos(
                df_isd, df_icd, uf_cols,
                mes_referencia, data_emissao
            )
            
            # Enviar preços (passando mes_referencia)
            log_precos, sucessos, falhas = enviar_precos_lote(precos, chaves, log_precos, mes_referencia)
            total_sucesso += sucessos
            total_falha += falhas
            
            print(f"  ✓ Arquivo concluído: {sucessos} sucessos, {falhas} falhas")
            
        except Exception as e:
            print(f"  ✗ Erro ao processar arquivo: {e}")
            import traceback
            traceback.print_exc()
            continue
    
    print("\n" + "=" * 70)
    print("RESUMO FINAL")
    print("=" * 70)
    print(f"✓ Total enviado com sucesso: {total_sucesso}")
    print(f"✗ Total com falha: {total_falha}")
    print(f"📊 Total processado: {total_sucesso + total_falha}")
    
    if total_falha > 0:
        print(f"\n⚠️  Verifique erros em: {ERRO_FILE}")
    
    print("=" * 70)

if __name__ == "__main__":
    main()